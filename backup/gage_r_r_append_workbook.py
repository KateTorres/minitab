# This file was created because I no longer have access to MiniTab software for free
# The content of this file re-creates template from training LinkedIn Training https://www.linkedin.com/learning/quality-analytics-using-minitab/ by Richard Chua  

    # Gage Repeatability and Reproucibility (Gage R&R) or Measurement System Analysis (MSA)
    # Minitab : Stat -> quality tool -> Gage Study -> Create Gage R&R Study worksheet
    # User enters # of operators, # of replicates, name of the operators.
    # Input is done via Python and recorded into the excel file 

# Make sure to install pandas openpyxl

# Necessary Imports
import os
import pandas as pd
from openpyxl import load_workbook

# Path and variable definitions
CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
FILE_NAME = 'gage_r_r_analysis.xlsx'
FILE_PATH = os.path.join(CURRENT_DIR, FILE_NAME)

# Load the data from the existing Excel file
def load_data(filepath):
    try:
        data = pd.read_excel(filepath)
        return data
    except Exception as e:
        print(f"An error occurred while loading the Excel file: {e}")
        return None

# Extract column name from Column D
def get_column_name(data):
    if data is not None:
        try:
            return data.columns[3]  # Get the name of the fourth column (index 3)
        except IndexError:
            print("Column D does not exist in the provided data.")
            return "Metric"
    return "Metric"

# Create a new worksheet in the existing workbook and add Gage R&R input form
def create_gage_rr_form(data, parts_count, workbook_path, column_name, operators, replicates):
    if data is None:
        return None
    wb = load_workbook(workbook_path)
    ws = wb.create_sheet("Gage R&R Input Form")

    # Create headers
    headers = ['Operator', 'RunOrder', 'Parts', column_name]
    ws.append(headers)
    
    # Generate form entries for the specified number of parts, operators, and replicates
    for operator in operators:
        for part in range(1, parts_count + 1):
            for replicate in range(1, replicates + 1):
                row = [operator, replicate, f'Part {part}', '']  # Placeholder for the metric measurement
                ws.append(row)

    return wb

# Main function to run the process
def main():
    data = load_data(FILE_PATH)
    if data is not None:
        column_name = get_column_name(data)  # Get dynamic column name from column D
        parts_count = int(input("Enter the number of parts: "))
        num_operators = int(input("Enter the number of operators: "))
        operators = [input(f"Enter the name of operator {i+1}: ") for i in range(num_operators)]
        replicates = int(input("Enter the number of replicates: "))

        wb = create_gage_rr_form(data, parts_count, FILE_PATH, column_name, operators, replicates)
        if wb is not None:
            wb.save(FILE_PATH)
            print(f"Updated workbook with 'Gage R&R Input Form' sheet saved to {FILE_PATH}")
        else:
            print("Failed to create the workbook. Please check the data.")
    else:
        print("Failed to load data. Please check the file and path.")

if __name__ == "__main__":
    main()