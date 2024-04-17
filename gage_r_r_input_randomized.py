# This file was created because I no longer have access to MiniTab software for free
# The content of this file re-creates template from training LinkedIn Training https://www.linkedin.com/learning/quality-analytics-using-minitab/ by Richard Chua  

    # Gage Repeatability and Reproucibility (Gage R&R) or Measurement System Analysis (MSA)
    # Minitab : Stat -> quality tool -> Gage Study -> Create Gage R&R Study worksheet
    # User enters # of operators, # of replicates, name of the operators.
    # Input is done via Python and recorded into the excel file 

# Make sure to install pandas openpyxl

# Necessary Imports
import os
import random
from openpyxl import Workbook, load_workbook  # Ensure this import statement is correct

# Path and variable definitions
CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
FILE_NAME = 'gage_r_r_input_randomized.xlsx'
FILE_PATH = os.path.join(CURRENT_DIR, FILE_NAME)

# Function to create or clear the workbook
def prepare_workbook(file_path):
    # Check if the file exists
    if os.path.exists(file_path):
        # Load the existing workbook
        wb = load_workbook(file_path)
        # Remove all existing sheets
        for sheet in wb.sheetnames:
            del wb[sheet]
    else:
        # Create a new workbook
        wb = Workbook()
    return wb

# Create a new Excel workbook and add Gage R&R input form
def create_gage_rr_form(parts_count, workbook_path, operators, replicates):
    wb = prepare_workbook(workbook_path)
    ws = wb.active
    if not ws:  # If no active sheet, create a new one
        ws = wb.create_sheet("Gage R&R Input Form")
    else:
        ws.title = "Gage R&R Input Form"

    # Define the headers for the worksheet
    headers = ['Operator', 'RunOrder', 'Parts', 'Metric']  # 'Metric' is the manual entry column
    ws.append(headers)
    
    # Generate form entries for the specified number of parts, operators, and replicates
    parts_list = [f'Part {i}' for i in range(1, parts_count + 1)]
    for operator in operators:
        for replicate in range(1, replicates + 1):
            random.shuffle(parts_list)  # Randomize the order of parts for each operator and replicate
            for part in parts_list:
                row = [operator, replicate, part, '']  # Placeholder for the metric measurement
                ws.append(row)

    # Save the workbook to the specified path
    wb.save(workbook_path)
    return wb

# Main function to run the process
def main():
    parts_count = int(input("Enter the number of parts: "))
    num_operators = int(input("Enter the number of operators: "))
    operators = [input(f"Enter the name of operator {i+1}: ") for i in range(num_operators)]
    replicates = int(input("Enter the number of replicates: "))

    wb = create_gage_rr_form(parts_count, FILE_PATH, operators, replicates)
    if wb is not None:
        print(f"New Gage R&R Input Form created and saved to {FILE_PATH}")
    else:
        print("Failed to create the workbook. Please check the data.")

if __name__ == "__main__":
    main()